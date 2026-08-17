"""Excel 指示文件解析器

解析亚马逊 FBA 派送表，处理合并单元格（前向填充 ffill），
提供箱唛 → FBA 信息 / SKU 列表 / 贴标顺序等数据查询接口。

支持三种格式：
  1. 标准指示文件（旧格式）：固定 8 列标题
  2. 客户原始派送表（新格式）：自动字段映射，份数 从 item no. 提取，SKU数量 从 総数量/总数量/SKU数量 列读取
  3. 单票贴标指示表：含"贴标操作要求"列，份数 从备注列提取，SKU数量 从 FNSKU总枚数 列读取
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any

# ─── 标准指示文件必需列（旧格式）────────────────────────────

REQUIRED_COLUMNS: list[str] = [
    "箱唛",
    "FBA番号",
    "箱数",
    "SKU",
    "SKU数量",
    "贴标顺序",
    "份数",
    "FBA仓库编码",
]


# ─── 派送表列名映射（新格式）────────────────────────────────

# 每个标准字段对应一组候选匹配模式（OR 关系），每组是 AND 关系的子串列表
DELIVERY_COLUMN_MAPPING: dict[str, list[list[str]]] = {
    "箱唛": [["箱唛号"], ["箱唛"], ["MARK"], ["mark"]],
    "FBA番号": [["FBA番号"], ["FBA编号"], ["FBA号"], ["FBA#"], ["FBA货箱编号"]],
    "箱数": [["箱数"], ["(CTNS)"], ["操作箱数"]],
    "SKU": [["SKU"], ["更新FNSKU"]],
    "SKU数量": [["総数量"], ["总数量"], ["SKU数量"]],  # 不用泛义"数量"列
    "贴标顺序": [["贴标顺序"], ["标签顺序"], ["对应的外箱标识"], ["外箱标识"]],
    "份数": [["份数"], ["每箱贴标"], ["打印份数"], ["Copies"]],  # 备选独立列
    "FBA仓库编码": [["FBA仓库"], ["FBA 仓"], ["仓库编码"], ["仓库代码"]],
    "对应SKU标识": [["对应SKU标识"], ["SKU标识"], ["对应SKU"]],  # 可选字段
    "FBA序列号": [["FBA序列号"]],
}

# 派送表必备字段（贴标顺序非强制，单票贴标指示表无此独立列）
DELIVERY_REQUIRED: list[str] = [
    "箱唛",
    "FBA番号",
    "箱数",
    "SKU",
    "FBA仓库编码",
]

# ─── 单票贴标指示表列名映射（第三种格式）─────────────────────────

# ── 单票贴标指示映射（第三种格式）───────────────────────────────────────────────

# 单票指示关键词：每行一个子模式（OR），子模式内 AND 关系子串匹配
# 支持 split header（两行表头），Row2 的子列名覆盖 Row1 的父列名
SINGLE_SHIPMENT_MAPPING: dict[str, list[list[str]]] = {
    "箱唛": [["YHD单号"], ["YHD系统单号"]],
    "贴标顺序": [["对应的外箱标识"]],
    "FBA番号": [["新", "FBA番号"], ["FBA番号"], ["FBA货箱编号"]],
    "箱数": [["操作的箱数"], ["操作箱数"]],
    "SKU": [["新", "FNSKU"], ["FNSKU"]],
    "SKU数量": [["FNSKU枚数"], ["FNSKU总枚数"]],
    "FBA仓库编码": [["新", "仓库代码"], ["仓库代码"]],
    "FBA序列号": [["FBA序列号"]],
    "备注": [["备注"], ["具体操作要求"], ["具体贴标操作要求"], ["第二批贴标操作要求"], ["操作要求"]],
}

# 单票指示必填字段
SINGLE_SHIPMENT_REQUIRED: list[str] = [
    "箱唛",
    "贴标顺序",
    "FBA番号",
    "箱数",
    "SKU",
    "SKU数量",
]

# ── 自定义列名映射（用户设置）───────────────────────────
# 必填 8 项：检测时必须全部命中才识别为 custom 格式
CUSTOM_REQUIRED: list[str] = [
    "箱唛",
    "FBA番号",
    "箱数",
    "SKU",
    "SKU数量",
    "贴标顺序",
    "FBA仓库编码",
    "FBA序列号",
]
# 可选 3 项：列存在则使用，不存在则容忍
CUSTOM_OPTIONAL: list[str] = [
    "对应SKU标识",
    "份数",
    "备注",
]

MAX_HEADER_ROW_SEARCH: int = 5


# ─── 数据容器 ────────────────────────────────────────────────


class ExcelData:
    """存储解析后的 Excel 数据，提供查询方法。"""

    def __init__(self, rows: list[dict[str, Any]]) -> None:
        self.rows = rows

    # ── 便捷列访问 ─────────────────────────────────────────────

    def column_values(self, name: str) -> list[Any]:
        return [row.get(name) for row in self.rows]

    def unique_values(self, name: str) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for row in self.rows:
            v = str(row.get(name, "")).strip()
            if v and v not in seen:
                seen.add(v)
                result.append(v)
        return result

    def non_none_values(self, name: str) -> list[Any]:
        result: list[Any] = []
        for row in self.rows:
            v = row.get(name)
            if v is not None and str(v).strip() != "":
                # 排除 NaN
                if isinstance(v, float) and math.isnan(v):
                    continue
                result.append(v)
        return result

    # ── 业务查询 ───────────────────────────────────────────────

    def get_fba_info(self, mark: str) -> dict[str, Any] | None:
        """根据箱唛查询 FBA 信息。"""
        mark = mark.strip()
        for row in self.rows:
            if str(row.get("箱唛", "")).strip() == mark:
                copies_raw = row.get("份数")
                try:
                    copies = int(float(copies_raw)) if copies_raw not in (None, "") else 0
                except (ValueError, TypeError):
                    copies = 0
                return {
                    "FBA番号": row["FBA番号"],
                    "份数": copies,
                    "贴标顺序": row.get("贴标顺序", ""),
                    "FBA仓库编码": row.get("FBA仓库编码") or "",
                    "item_no": str(row.get("item_no", "") or "").strip(),
                }
        return None

    def get_all_sticker_orders(self, mark: str) -> list[str]:
        """获取指定箱唛的所有贴标顺序（去重），提取 'U' 之后的部分。"""
        mark = mark.strip()
        seen: set[str] = set()
        orders: list[str] = []
        for row in self.rows:
            if str(row.get("箱唛", "")).strip() == mark:
                order = str(row.get("贴标顺序", ""))
                if order and order not in seen:
                    seen.add(order)
                    orders.append(order)

        # 提取 "U" 之后的数字部分，去除前导零，保留负号-后的数字也去除前导零
        processed: list[str] = []
        for o in orders:
            processed.append(extract_tail_number(o))
        return processed


    def get_sticker_orders_with_fba(self, mark: str) -> dict:
        """获取指定箱唛的贴标顺序和 FBA 序列号，检查尾号是否一致。

        Returns:
            {
                "all_match": bool,
                "items": [{"sticker": "1-25", "fba_seq": "FBA15GD26922U000001-25"}, ...]
            }
        """
        import re
        mark = mark.strip()
        seen_sticker: set[str] = set()
        items: list[dict] = []
        all_match = True

        for row in self.rows:
            if str(row.get("箱唛", "")).strip() != mark:
                continue
            raw_order = str(row.get("贴标顺序", "")).strip()
            if not raw_order or raw_order in seen_sticker:
                continue
            seen_sticker.add(raw_order)

            # 获取 FBA 序列号
            fba_seq = str(row.get("FBA序列号", "")).strip()

            # 提取 U 后的数字部分（复用公共 helper）
            order_num = extract_tail_number(raw_order)
            fba_num = extract_tail_number(fba_seq) if fba_seq else ""

            if fba_num and order_num != fba_num:
                all_match = False

            items.append({
                "sticker": order_num,
                "fba_seq": fba_seq,
                "sticker_raw": raw_order,
            })

        return {"all_match": all_match, "items": items}
    def get_box_count_for_mark(self, mark: str) -> int:
        """获取指定箱唛的总箱数。"""
        mark = mark.strip()
        total = 0.0
        for row in self.rows:
            if str(row.get("箱唛", "")).strip() == mark:
                val = row.get("箱数")
                if val not in (None, ""):
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
        return int(total)

    def get_skus_for_mark(self, mark: str) -> list[dict[str, Any]]:
        """获取指定箱唛下的所有 SKU 信息。

        返回字典包含:
          - SKU, SKU数量, 贴标顺序
          - 对应SKU标识 (原始值), SKU标识符 (去 -IDn 后缀)
          - 复合键 (用于 PDF 页面匹配和数据库查找)
        """
        mark = mark.strip()
        results: list[dict[str, Any]] = []
        for row in self.rows:
            if str(row.get("箱唛", "")).strip() == mark:
                sku = row.get("SKU")
                if sku is not None and str(sku).strip() != "":
                    qty_raw = row.get("SKU数量")
                    try:
                        qty = int(float(qty_raw)) if qty_raw not in (None, "") else 0
                    except (ValueError, TypeError):
                        qty = 0
                    raw_identifier = str(row.get("对应SKU标识", "")).strip()
                    identifier = strip_sku_identifier(raw_identifier)
                    results.append(
                        {
                            "SKU": str(sku),
                            "SKU数量": qty,
                            "贴标顺序": row.get("贴标顺序", ""),
                            "对应SKU标识": raw_identifier,
                            "SKU标识符": identifier,
                            "复合键": make_compound_sku_key(str(sku), identifier),
                        }
                    )
        return results

    def get_all_compound_sku_info(self) -> list[dict[str, str]]:
        """返回所有 SKU 的复合键信息，用于 PDF 页面匹配。

        Returns:
            [{"sku": ..., "identifier": ..., "compound_key": ...}, ...]
            去重，同一复合键只出现一次。
        """
        seen: set[str] = set()
        result: list[dict[str, str]] = []
        for row in self.rows:
            sku = str(row.get("SKU", "")).strip()
            if not sku:
                continue
            raw_identifier = str(row.get("对应SKU标识", "")).strip()
            identifier = strip_sku_identifier(raw_identifier)
            compound_key = make_compound_sku_key(sku, identifier)
            if compound_key not in seen:
                seen.add(compound_key)
                result.append(
                    {
                        "sku": sku,
                        "identifier": identifier,
                        "compound_key": compound_key,
                    }
                )
        return result

    def search_sku(self, query: str) -> list[str]:
        """模糊搜索 SKU。"""
        query = query.lower()
        matches: list[str] = []
        seen: set[str] = set()
        for row in self.rows:
            sku = str(row.get("SKU", ""))
            if sku and query in sku.lower():
                if sku not in seen:
                    seen.add(sku)
                    matches.append(sku)
        return matches

    @property
    def box_marks(self) -> list[str]:
        """返回去重保持顺序的箱唛列表。"""
        seen: set[str] = set()
        result: list[str] = []
        for row in self.rows:
            mark = str(row.get("箱唛", "")).strip()
            if mark and mark not in seen:
                seen.add(mark)
                result.append(mark)
        return result


# ─── 份数 / SKU 数量提取 ─────────────────────────────────────

# ─── 中文数字转换 ────────────────────────────────────────────────

_CN_DIGIT_MAP: dict[str, int] = {
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}

# 用于正则匹配：阿拉伯数字 或 中文数字（一～九、两）
_CN_NUM_PAT = r"[一二两三四五六七八九\d]+"


def _parse_num(raw: str) -> int:
    """将含中文数字的字符串转为整数。"""
    if raw.isdigit():
        return int(raw)
    # 中文数字逐字转换
    total = 0
    for ch in raw:
        total = total * 10 + _CN_DIGIT_MAP.get(ch, 0)
    return total


# ─── 份数 / SKU 数量提取 ─────────────────────────────────────


def extract_copies_from_item_no(text: str | None) -> tuple[int, str]:
    """从 item no. / 备注列文本中提取每箱打印份数。

    匹配模式（支持阿拉伯数字 + 中文数字）：
      - "FBA每箱贴 N 张"   → 份数 = N
      - "每箱贴 N 张"       → 份数 = N
      - 无匹配时返回 0，由调用方默认为 1 张
        （不回退取最后一个数字，避免把单号/日期等误当份数）

    Returns:
        (份数, 来源描述)
    """
    if not text:
        return 0, "无 item no. 信息"

    s = str(text)

    # 模式 1: FBA每箱贴N张（最可靠）
    m = re.search(rf"FBA\s*每箱\s*[贴貼]\s*({_CN_NUM_PAT})\s*张", s)
    if m:
        n = _parse_num(m.group(1))
        return n, f"item no. → FBA每箱贴{n}张"

    # 模式 2: 每箱贴N张
    m = re.search(rf"每箱\s*[贴貼]\s*({_CN_NUM_PAT})\s*张", s)
    if m:
        n = _parse_num(m.group(1))
        return n, f"item no. → 每箱贴{n}张"

    # 未找到明确的"贴N张"信息：返回 0，由调用方默认为 1 张
    return 0, "未找到FBA贴N张，份数默认为1"


# ─── FBA 编号清洗 ────────────────────────────────────────────────


def clean_fba_number(raw: str | None) -> str:
    """从可能含后缀的 FBA 编号中提取纯编号。

    示例：
      "FBA15GCK8TVJ"          → "FBA15GCK8TVJ"
      "FBA15GCK8TVJ U01-25"   → "FBA15GCK8TVJ"
      "FBA15GCK8TVJ (U01-25)" → "FBA15GCK8TVJ"
      ""                      → ""

    规则：取 FBA 前缀开始到首个空格/括号/换行 之前的连续文本。
    如果值不以 FBA 开头，原样返回。
    """
    if not raw:
        return ""
    s = str(raw).strip()
    if not s.upper().startswith("FBA"):
        return s
    # 取 FBA 之后的连续字母数字到首个空白/分隔符
    m = re.match(r"(FBA\w+)", s)
    return m.group(1) if m else s


# ─── SKU 标识符工具 ────────────────────────────────────────────


def strip_sku_identifier(raw_id: str) -> str:
    """从"对应SKU标识"中移除 ID 字样，保留数字后缀。

    例如: "5261614348094-ID2" → "5261614348094-2"
    如果为空字符串或无可识别后缀，原样返回。
    """
    if not raw_id:
        return ""
    return re.sub(r"-ID(\d+)$", r"-\1", str(raw_id).strip())


def make_compound_sku_key(sku: str, identifier: str) -> str:
    """生成复合 SKU 键，用于区分同一 SKU 的不同标识。

    格式: "SKU|||标识符" 或纯 "SKU"（无标识符时）。
    """
    if identifier:
        return f"{sku}|||{identifier}"
    return sku


# ─── 列名模糊匹配 ────────────────────────────────────────────


def _normalize_header(header: str) -> str:
    """规范化列名：将换行符等空白字符替换为单个空格，便于匹配。"""
    if not header:
        return ""
    return re.sub(r"\s+", " ", str(header)).strip()


def extract_tail_number(val: str) -> str:
    """从贴标顺序 / FBA序列号中提取 "U" 之后的数字部分。

    兼容多种情况：
      - 序列号中重复出现 U（取最后一个 U 之后的内容）
      - 分隔符支持短横 -、波浪 ~、中文顿号 、、中文逗号 ，
      - 单个分段不是纯数字时保留原样，不抛异常
    """
    if not val:
        return ""
    idx = -1
    for ch in ("U", "u"):
        i = val.rfind(ch)
        if i > idx:
            idx = i
    if idx < 0:
        return val
    after_u = val[idx + 1:]
    parts = re.split(r"[-~\u3001\uff0c]", after_u)
    clean_parts: list[str] = []
    for p in parts:
        try:
            clean_parts.append(str(int(p)))
        except ValueError:
            clean_parts.append(p)
    return "-".join(clean_parts)


def _match_columns(
    headers: list[str],
    mapping: dict[str, list[list[str]]],
) -> dict[str, int]:
    """将标题行按映射规则匹配到标准字段列索引。

    Args:
        headers: 标题行（已 strip）。
        mapping: {标准字段: [[候选子串 AND 组], ...]}。

    Returns:
        {标准字段: 列索引} 字典。未匹配的字段不在其中。

    匹配策略（修复：SKU 误配到"对应SKU标识"列）：
      1. 第一轮：候选词与表头"完全相等"才命中（如 "SKU" 只匹配列名 "SKU"）。
      2. 第二轮：全部列精确匹配失败后，回退到"子串包含"匹配，兼容旧文件。
    """
    col_idx: dict[str, int] = {}

    def _scan(sub_match: bool) -> None:
        for std_name, candidates in mapping.items():
            if std_name in col_idx:
                continue
            for col_i, header in enumerate(headers):
                if not header:
                    continue
                # 规范化列名（合并换行符等空白字符）
                normalized = _normalize_header(header)
                for candidate in candidates:
                    subs = [_normalize_header(sub) for sub in candidate]
                    if sub_match:
                        hit = all(s in normalized for s in subs)
                    else:
                        hit = all(s == normalized for s in subs)
                    if hit:
                        col_idx[std_name] = col_i
                        break
                if std_name in col_idx:
                    break

    _scan(sub_match=False)  # 精确匹配优先
    _scan(sub_match=True)   # 子串包含兜底
    return col_idx


# ─── 格式检测器 ──────────────────────────────────────────────


def _match_custom_columns(
    headers: list[str],
    mapping: dict[str, str],
    match_mode: str = "contains",
) -> dict[str, int]:
    """按用户配置的关键字匹配列索引。

    Args:
        headers: 表头行（已 strip）。
        mapping: {标准字段: 列名关键字}。
        match_mode: "contains" 包含关键字（不区分大小写） / "exact" 完全等于。

    Returns:
        {标准字段: 列索引}，未匹配的字段不在其中。
    """
    col_idx: dict[str, int] = {}
    for std_name, keyword in mapping.items():
        if not keyword:
            continue
        kw = str(keyword).strip()
        if not kw:
            continue
        for col_i, header in enumerate(headers):
            if not header:
                continue
            nh = _normalize_header(header)
            if match_mode == "exact":
                if nh == kw:
                    col_idx[std_name] = col_i
                    break
            else:
                if kw.lower() in nh.lower():
                    col_idx[std_name] = col_i
                    break
    return col_idx

class FieldMapper:
    """智能格式检测与字段映射。

    支持三种格式：
      - standard: 标准指示文件（8 列精确匹配）
      - delivery: 客户原始派送表（模糊匹配 + item no. 提取）
      - label_instruction: 单票贴标指示表（专用列映射 + 备注列提取）
    """

    @staticmethod
    def detect(
        headers: list[str],
        custom_mapping: dict[str, str] | None = None,
        match_mode: str = "contains",
        custom_priority: bool = False,
    ) -> dict[str, Any] | None:
        """检测格式类型并返回映射信息。

        Args:
            headers: 标题行（已 strip）。
            custom_mapping: 用户自定义 {标准字段: 列名关键字}。
            match_mode: "contains" / "exact"。
            custom_priority: 为 True 时优先尝试自定义映射。

        Returns:
            None → 无法识别此格式
            dict → {"format", "col_idx", "copies_source", "sku_qty_source", "warnings"}
        """
        # 0) 自定义列名映射（用户设置，勾选"优先自定义"时优先）
        if custom_priority and custom_mapping:
            custom_col_idx = _match_custom_columns(headers, custom_mapping, match_mode)
            custom_missing = [
                c for c in CUSTOM_REQUIRED if c not in custom_col_idx
            ]
            if not custom_missing:
                # 可选字段：能匹配则加入 col_idx，否则容忍
                for opt in CUSTOM_OPTIONAL:
                    if opt not in custom_col_idx and opt in custom_mapping:
                        opt_idx = _match_custom_columns(
                            headers, {opt: custom_mapping[opt]}, match_mode
                        )
                        if opt in opt_idx:
                            custom_col_idx[opt] = opt_idx[opt]
                return {
                    "format": "custom",
                    "col_idx": custom_col_idx,
                    "copies_source": "custom",
                    "sku_qty_source": "column",
                    "warnings": [],
                }

        # 1) 标准格式精确匹配
        if all(c in headers for c in REQUIRED_COLUMNS):
            return {
                "format": "standard",
                "col_idx": {c: headers.index(c) for c in REQUIRED_COLUMNS},
                "copies_source": "column",
                "sku_qty_source": "column",
                "warnings": [],
            }

        # 2) 单票贴标指示表优先检测（第三种格式——有"贴标操作要求"或"更新FNSKU"等特征列）
        label_col_idx = _match_columns(headers, SINGLE_SHIPMENT_MAPPING)
        label_missing = [c for c in SINGLE_SHIPMENT_REQUIRED if c not in label_col_idx]
        if not label_missing:
            # 确认有备注类列（贴标操作要求 / 备注 / 说明 等）
            has_notes = any(
                _normalize_header(h)
                and any(
                    kw in _normalize_header(h)
                    for kw in ["贴标操作要求", "贴标要求", "备注", "说明", "操作要求"]
                )
                for h in headers
            )
            if has_notes:
                # ── 新字段优先：当既有原列又有新列时，优选新列 ──
                prefer_new_cols = {}
                for ci, h in enumerate(headers):
                    nh = _normalize_header(h)
                    if nh == "新FBA番号" or nh.startswith("新FBA番号"):
                        prefer_new_cols["FBA番号"] = ci
                    if nh == "新FNSKU" or nh.startswith("新FNSKU"):
                        prefer_new_cols["SKU"] = ci
                    if nh == "新仓库代码" or nh.startswith("新仓库代码"):
                        prefer_new_cols["FBA仓库编码"] = ci
                    if "系统单号" in nh:
                        prefer_new_cols["箱唛"] = ci
                for field, col in prefer_new_cols.items():
                    if col is not None:
                        label_col_idx[field] = col
                return {
                    "format": "shipment_instruction",
                    "col_idx": label_col_idx,
                    "copies_source": "item_no",
                    "sku_qty_source": "column",
                    "warnings": [
                        "单票贴标指示表 — 份数从贴标操作要求列提取，SKU数量从FNSKU总枚数列取",
                    ],
                }

        # 3) 派送表格式模糊匹配
        col_idx = _match_columns(headers, DELIVERY_COLUMN_MAPPING)

        # 检查必备字段
        missing = [c for c in DELIVERY_REQUIRED if c not in col_idx]
        if missing:
            return None  # 无法识别

        # 确定份数和 SKU 数量的来源
        copies_source = "column" if "份数" in col_idx else "item_no"
        # SKU数量 仅从列读取；无 SKU数量 列时固定默认 1（不从 item no. 提取）
        sku_qty_source = "column"

        warnings: list[str] = []
        if copies_source == "item_no":
            warnings.append("份数将自动从 item no. 列提取（'FBA每箱贴N张'）")

        return {
            "format": "delivery",
            "col_idx": col_idx,
            "copies_source": copies_source,
            "sku_qty_source": sku_qty_source,
            "warnings": warnings,
        }


# ─── 解析器 ──────────────────────────────────────────────────


class ExcelLoader:
    """读取 Excel 指示文件，处理合并单元格（ffill 前向填充）。

    自动检测格式：标准指示文件 或 客户原始派送表。
    """

    def parse(
        self,
        file_path: str | Path,
        custom_mapping: dict[str, str] | None = None,
        match_mode: str = "contains",
        custom_priority: bool = False,
    ) -> tuple[ExcelData | None, str]:
        """解析 Excel 文件，返回 (ExcelData | None, 消息)。

        Args:
            file_path: Excel 文件路径。
            custom_mapping: 用户自定义 {标准字段: 列名关键字}。
            match_mode: "contains" / "exact"。
            custom_priority: 为 True 时优先尝试自定义映射。
        """
        file_path = Path(file_path)
        if not file_path.exists():
            return None, "文件未找到。"

        try:
            rows_raw = self._read_excel(file_path)
            if not rows_raw:
                return None, "Excel 文件为空。"

            # ── 标题行检测 ──
            result = self._detect_header(
                rows_raw,
                custom_mapping=custom_mapping,
                match_mode=match_mode,
                custom_priority=custom_priority,
            )
            if result is None:
                return None, "无法识别 Excel 格式，请确认文件包含正确的表头。"
            header_idx, headers, mapping = result

            if mapping is None:
                # 生成诊断信息
                found_cols = [
                    h
                    for h in headers
                    if h and not h.startswith("柜号") and not h.startswith("TOTAL")
                ][:15]
                return None, (
                f"无法识别 Excel 格式。请确认文件为：\n"
                f"  1) 标准指示文件（含列: {', '.join(REQUIRED_COLUMNS[:4])}...）\n"
                f"  2) 客户派送表（含列: 送货单号, FBA编号, SKU 等）\n"
                f"  3) 单票贴标指示表（含列: YHD单号, 外箱标识, 新FBA番号, FNSKU 等）\n"
                    f"检测到的前 {len(found_cols)} 个列: {', '.join(found_cols)}"
                )

            fmt = mapping["format"]
            col_idx = mapping["col_idx"]

            # ── 构建格式描述 ──
            if fmt == "standard":
                fmt_desc = "标准指示文件（旧格式）"
            elif fmt == "shipment_instruction":
                fmt_desc = "单票贴标指示表（第三种格式）"
            elif fmt == "custom":
                fmt_desc = "自定义列名格式"
            else:
                fmt_desc = "派送表（新格式）"

            # ── 解析数据行 ──
            data_rows = self._parse_rows(
                rows_raw,
                col_idx,
                headers,
                header_idx,
                copies_source=mapping["copies_source"],
                sku_qty_source=mapping["sku_qty_source"],
            )

            warnings = mapping.get("warnings", [])
            msg_parts = [f"{fmt_desc} — 加载成功，{len(data_rows)} 行数据"]
            if warnings:
                msg_parts.append(" | " + "；".join(warnings))

            return ExcelData(data_rows), " ".join(msg_parts)

        except Exception as e:
            return None, f"解析 Excel 出错: {e}"

    # ── 标题行检测 ───────────────────────────────────────────

    @staticmethod
    def _detect_header(
        rows_raw: list[list[Any]],
        custom_mapping: dict[str, str] | None = None,
        match_mode: str = "contains",
        custom_priority: bool = False,
    ) -> tuple[int, list[str], dict[str, Any] | None]:
        """在前 N 行中寻找标题行。

        Args:
            custom_mapping: 用户自定义列名映射。
            match_mode: "contains" / "exact"。
            custom_priority: 为 True 时优先尝试自定义映射。

        Returns:
            (header_row_index, headers, mapping_or_None)
        """
        for i in range(min(MAX_HEADER_ROW_SEARCH, len(rows_raw))):
            headers = [str(h).strip() if h is not None else "" for h in rows_raw[i]]
            # 跳过几乎全空的行
            non_empty = sum(1 for h in headers if h)
            if non_empty < 3:
                continue
            # ── Split-header 检测：检查下一行是否有子列名 ──
            # 当下一行也有非空单元格时，合并两行（Row2 覆盖 Row1 的非空值）
            if i + 1 < len(rows_raw):
                next_row = rows_raw[i + 1]
                next_non_empty = sum(1 for h in next_row if h is not None and str(h).strip())
                if next_non_empty >= 1:
                    row2 = [str(h).strip() if h is not None else "" for h in next_row]
                    # 补齐列数
                    while len(row2) < len(headers):
                        row2.append("")
                    merged = headers[:]
                    for ci in range(min(len(headers), len(row2))):
                        if row2[ci]:
                            merged[ci] = row2[ci]
                    # 用合并后的表头尝试匹配
                    mapping = FieldMapper.detect(
                        merged,
                        custom_mapping=custom_mapping,
                        match_mode=match_mode,
                        custom_priority=custom_priority,
                    )
                    if mapping is not None:
                        # Remove sub-header row from raw data so parsing starts from correct row
                        rows_raw.pop(i + 1)
                        return i, merged, mapping
            # ── 单行表头匹配 ──
            mapping = FieldMapper.detect(
                        headers,
                        custom_mapping=custom_mapping,
                        match_mode=match_mode,
                        custom_priority=custom_priority,
                    )
            if mapping is not None:
                return i, headers, mapping

        # No matching header found
        return None

    # ── 内部方法 ───────────────────────────────────────────────

    @staticmethod
    def _read_excel(path: Path) -> list[list[Any]]:
        suffix = path.suffix.lower()
        if suffix == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(path))
            sheet = wb.sheet_by_index(0)
            return [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)
            ]
        else:
            from openpyxl import load_workbook

            wb = load_workbook(str(path), read_only=True, data_only=True)
            try:
                sheet = wb.active
                return [list(row) for row in sheet.iter_rows(values_only=True)]
            finally:
                wb.close()

    @staticmethod
    def _parse_rows(
        rows_raw: list[list[Any]],
        col_idx: dict[str, int],
        headers: list[str],
        header_idx: int,
        copies_source: str = "column",
        sku_qty_source: str = "column",
    ) -> list[dict[str, Any]]:
        """逐行解析，对合并单元格执行前向填充。

        Args:
            rows_raw: 原始行数据（含标题行）。
            col_idx: {标准字段: 列索引}。
            headers: 标题行（已 strip）。
            header_idx: 标题行所在索引。
            copies_source: "column" 或 "item_no"。
            sku_qty_source: "column" 或 "item_no"。
        """
        # 确定哪些列需要 ffill（合并单元格列）
        ffill_columns: list[str] = ["箱唛", "FBA番号", "FBA仓库编码", "贴标顺序"]
        if copies_source == "column":
            ffill_columns.append("份数")
        if sku_qty_source == "column":
            ffill_columns.append("SKU数量")

        # 份数和 SKU 数量对应的列索引（如果来源于 item no.）
        item_no_col: int | None = None
        for candidate in ["item no.", "Item No.", "ITEM", "item"]:
            if candidate in headers:
                item_no_col = headers.index(candidate)
                break

        # 派送指示表：优先找备注 / 贴标操作要求 列（比 SKU+1 回退更可靠）
        notes_col: int | None = None
        if item_no_col is None:
            for candidate in [
                "贴标操作要求",
                "贴标要求",
                "备注",
                "说明",
                "具体要求",
                "操作要求",
            ]:
                if candidate in headers:
                    notes_col = headers.index(candidate)
                    break

        # 如果没有独立 item no. 列，尝试用 SKU 列附近的列（派送表通常 SKU 在第17列，item no. 在第18列）
        # 回退：用映射的备注列位置
        if notes_col is None and item_no_col is None and copies_source == "item_no":
            beikey = col_idx.get("备注")
            if beikey is not None:
                notes_col = beikey
        if item_no_col is None and notes_col is None and copies_source == "item_no":
            sku_col = col_idx.get("SKU")
            if sku_col is not None and sku_col + 1 < len(headers):
                item_no_col = sku_col + 1
        # 自定义格式：备注列来自自定义映射
        if notes_col is None and "备注" in col_idx:
            notes_col = col_idx["备注"]

        last: dict[str, Any] = {}
        for col in ffill_columns:
            if col == "份数":
                last[col] = 0
            elif col == "SKU数量":
                # 无 SKU数量 列时默认 0（打印时红色警报跳过）；有列时空单元格沿用前值或 1
                last[col] = 1 if "SKU数量" in col_idx else 0
            else:
                last[col] = ""

        parsed: list[dict[str, Any]] = []

        for r_idx in range(header_idx + 1, len(rows_raw)):
            cells = list(rows_raw[r_idx])
            # 补齐列
            if len(cells) < len(headers):
                cells.extend([None] * (len(headers) - len(cells)))

            # 取标准字段的值
            item: dict[str, Any] = {c: cells[col_idx[c]] for c in col_idx}

            # ── 捕获未映射列的值（供供防火墙等下游扫描）──
            mapped_indices = set(col_idx.values())
            for ci, cv in enumerate(cells):
                if ci in mapped_indices:
                    continue
                if cv is not None and str(cv).strip():
                    item[f"_extra_{ci}"] = str(cv).strip()

            # ── FBA 编号清洗（去掉 "U01-25" 等后缀）──
            fba_raw = item.get("FBA番号")
            if fba_raw is not None:
                item["FBA番号"] = clean_fba_number(str(fba_raw))

            # 跳过 SKU 为空的行（汇总行 / 无 SKU 行）
            sku_raw = item.get("SKU")
            if sku_raw is None or str(sku_raw).strip() == "":
                continue

            # 跳过合计/小计/总计等汇总行
            sku_str = str(sku_raw).strip()
            if sku_str in ("合计", "小计", "总计", "合計", "小計", "総計"):
                continue

            # ── 前向填充 ──
            # 当箱唛切换为新值时，重置贴标顺序，防止跨箱唛泄漏
            _raw_box = item.get("箱唛")
            current_mark = str(_raw_box).strip() if _raw_box is not None else ""
            if current_mark and current_mark != str(last.get("箱唛", "")).strip():
                last["贴标顺序"] = ""
                if copies_source == "item_no":
                    # 份数按箱唛重置，防止上一箱唛的份数泄漏到无份数信息的新箱唛
                    last["份数"] = 0
            for col in ffill_columns:
                val = item.get(col)
                if val is not None and str(val).strip() != "":
                    last[col] = str(val).strip() if col != "份数" else val
                item[col] = last[col]

            # ── 备注列原始值（供防火墙等下游使用）──
            if notes_col is not None and notes_col < len(cells):
                raw_notes = cells[notes_col]
                if raw_notes is not None and str(raw_notes).strip() != "":
                    item["备注"] = str(raw_notes).strip()

            # ── item no. 原始值（供 FBA 备注打印使用）──
            if item_no_col is not None and item_no_col < len(cells):
                raw_item_no = cells[item_no_col]
                if raw_item_no is not None and str(raw_item_no).strip() != "":
                    item["item_no"] = str(raw_item_no).strip()
            if "item_no" not in item:
                item["item_no"] = str(item.get("备注", "") or "").strip()

            # ── 份数提取（从 item no. / 备注列）──
            if copies_source == "item_no":
                raw = None
                if item_no_col is not None and item_no_col < len(cells):
                    raw = cells[item_no_col]
                elif notes_col is not None and notes_col < len(cells):
                    raw = cells[notes_col]
                copies = 0
                if raw is not None and str(raw).strip() != "":
                    copies, _ = extract_copies_from_item_no(raw)
                if copies > 0:
                    last["份数"] = copies
                elif last.get("份数", 0) <= 0:
                    # 指示中没有份数，或备注中没有找到 FBA贴N张：默认为 1 张
                    last["份数"] = 1
                item["份数"] = last["份数"]

            # ── 份数提取（自定义格式：份数列 → 备注列 → 默认 1）──
            if copies_source == "custom":
                copies = 0
                if "份数" in col_idx:
                    raw = item.get("份数")
                    try:
                        copies = int(float(raw)) if raw not in (None, "") else 0
                    except (ValueError, TypeError):
                        copies = 0
                if copies <= 0 and notes_col is not None and notes_col < len(cells):
                    raw_notes = cells[notes_col]
                    if raw_notes is not None and str(raw_notes).strip():
                        copies, _ = extract_copies_from_item_no(raw_notes)
                if copies <= 0:
                    copies = 1
                item["份数"] = copies

            parsed.append(item)

        return parsed
