# -*- coding: utf-8 -*-
"""
FNSKU和商品描述提取工具 V5
从FBA标签PDF中提取FNSKU及其下方的商品描述。
逐页独立处理：文本页提取FNSKU，图片页转为PNG — 混合内容PDF不再遗漏。
输出为 .xlsx 格式，支持 Excel 直接打开。可选自动生成 50×30mm 条形码标签 PDF。

⚠️ 此为 V5 稳定版本，非特殊指令不得修改任何逻辑。
"""

import argparse
import os
import re
import sys

import fitz  # PyMuPDF
fitz.TOOLS.mupdf_display_errors(False)  # 抑制非关键 PDF 语法警告

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# OCR 支持（扫描件 PDF 识别）
_OCR_AVAILABLE = False
try:
    import pytesseract
    from PIL import Image

    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    # 优先用项目自带 tessdata（经过验证的日文模型），回退 Tesseract 默认
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _project_tessdata = os.path.join(_script_dir, "..", "..", "tessdata")
    if os.path.isdir(_project_tessdata):
        os.environ["TESSDATA_PREFIX"] = os.path.abspath(_project_tessdata)
    else:
        _default_tessdata = os.path.join(
            os.path.dirname(pytesseract.pytesseract.tesseract_cmd), "tessdata"
        )
        if os.path.isdir(_default_tessdata):
            os.environ["TESSDATA_PREFIX"] = _default_tessdata
    # 验证 Tesseract 可用
    import subprocess

    r = subprocess.run(
        [pytesseract.pytesseract.tesseract_cmd, "--version"],
        capture_output=True,
        text=True,
        timeout=5,
    )
    _OCR_AVAILABLE = r.returncode == 0
except Exception:
    pass

# 10-13位字母数字组合，二次过滤只保留疑似FNSKU
_FNSKU_RE = re.compile(r"[A-Z0-9]{10,13}")
# 宽松匹配：容忍字符间插入空格（如 "X00 1C40HN3"），带边界防止与相邻文字粘连
_FNSKU_RE_LOOSE = re.compile(
    r"(?<![A-Z0-9])X\s*0\s*0\s*[A-Z0-9](?:\s*[A-Z0-9]){6,9}?(?![A-Z0-9])"
)
# 已知的Amazon FNSKU常见前缀
_FNSKU_PREFIXES = ("X00", "B0")


def _loose_fnsku_matches(text):
    """逐行宽松匹配 FNSKU（容忍字符间插入空格），返回去空格后的编码列表。

    逐行匹配可避免 FNSKU 与相邻行/相邻词粘连成更长的字母数字串。
    """
    results: list[str] = []
    for line in text.splitlines():
        for m in _FNSKU_RE_LOOSE.findall(line):
            code = re.sub(r"\s+", "", m)
            if code not in results:
                results.append(code)
    return results


def _clean_desc(desc):
    """清理描述中的无意义字符串（如 TCPDF 签名），并将带圈数字转为普通数字"""
    if not desc:
        return desc
    import re

    # 移除控制字符（OCR 可能引入 \x00-\x1F 等非法字符）
    desc = re.sub("[\x00-\x1f\x7f-\x9f]", "", desc)

    # 带圈数字 → 普通数字（OCR 常把 12345 误识别为 ①②③④⑤）
    _CIRCLED = {
        "①": "1",
        "②": "2",
        "③": "3",
        "④": "4",
        "⑤": "5",
        "⑥": "6",
        "⑦": "7",
        "⑧": "8",
        "⑨": "9",
        "⑩": "10",
        "⑪": "11",
        "⑫": "12",
        "⑬": "13",
        "⑭": "14",
        "⑮": "15",
        "⑯": "16",
        "⑰": "17",
        "⑱": "18",
        "⑲": "20",
        "⓪": "0",
    }
    for circled, digit in _CIRCLED.items():
        desc = desc.replace(circled, digit)

    # 移除 CJK 字符之间的多余空格（OCR 常在每个字符间加空格）
    cjk = "[\u4e00-\u9fff\u3040-\u309f\u30a0-\u30ff]"
    pattern = "(" + cjk + ") +(" + cjk + ")"
    prev = None
    while prev != desc:
        prev = desc
        desc = re.sub(pattern, "\\1\\2", desc)

    patterns = [
        r"Powered\s+by\s+TCPDF.*?org[/\w]*",
        r"www\.tcpdf\.org[/\w]*",
        r"tcpdf\.org[/\w]*",
        r"Powered\s+by\s+TCPDF",
    ]
    cleaned = desc
    for p in patterns:
        cleaned = re.sub(p, "", cleaned, flags=re.IGNORECASE)
    cleaned = " ".join(cleaned.split())
    return cleaned.strip()


def _is_likely_fnsku(code):
    """二次校验：匹配到的编码是否像真实FNSKU"""
    return code.startswith(_FNSKU_PREFIXES)


def _extract_description(lines, start_idx):
    """从FNSKU所在行之后提取多行商品描述。

    收集直到遇到下一个FNSKU或条码行，最多向前看6行。
    跳过中间的空行，自动过滤 Powered by TCPDF 等无意义字符串。
    """
    desc_parts = []
    end = min(start_idx + 7, len(lines))
    for j in range(start_idx + 1, end):
        line = lines[j].strip()
        if not line:
            continue  # 跳过空行
        # 碰到下一个FNSKU或条码行则停止（FNSKU 行可能带空格，如 "X00 1C40HN3"）
        if (
            _FNSKU_RE.fullmatch(line)
            or _FNSKU_RE.fullmatch(re.sub(r"\s+", "", line))
            or re.match(r"^\*[\d ]+\*$", line)
        ):
            break
        # 过滤 TCPDF 签名
        if "tcpdf" in line.lower():
            continue
        desc_parts.append(line)
    return " ".join(desc_parts) if desc_parts else ""


def find_sku_files(source_dir):
    """递归搜索符合条件的PDF（文件名含X00/B0且不含FBA）"""
    files = []
    for root, _dirs, fns in os.walk(source_dir):
        for fn in fns:
            if not fn.lower().endswith(".pdf"):
                continue
            full_path = os.path.join(root, fn)
            # 跳过脚本自己的输出文件
            if "fnsku_result" in fn or fn.startswith("~$"):
                continue
            if ("X00" in fn or "B0" in fn) and "FBA" not in fn:
                files.append(full_path)
    return sorted(files)


def _extract_fnsku_from_filename(filename):
    """从文件名开头提取 FNSKU（如 X0013S61M9-扫描件.pdf → X0013S61M9）"""
    basename = os.path.splitext(os.path.basename(filename))[0]
    m = re.match(r"([A-Z0-9]{10,13})", basename)
    if m and _is_likely_fnsku(m.group(1)):
        return m.group(1)
    return None


def _extract_fnsku_from_ocr(text):
    """从 OCR 文本中提取 FNSKU（容忍字符间插入空格）"""
    matches = _FNSKU_RE.findall(text)
    if not matches:
        matches = _loose_fnsku_matches(text)
    for code in matches:
        if _is_likely_fnsku(code):
            return code
    return None


def _try_ocr_page(png_path, filename):
    """对单张 PNG 图片运行 OCR，提取 FNSKU 和 Description。

    优先从文件名提取 FNSKU（更准确），OCR 仅提取描述。
    返回: (fnsku, description) 或 (None, None)
    """
    global _OCR_AVAILABLE
    if not _OCR_AVAILABLE:
        return None, None
    try:
        # 优先从文件名提取 FNSKU（比 OCR 更准确）
        fnsku = _extract_fnsku_from_filename(filename)
        if not fnsku:
            return None, None

        text = pytesseract.image_to_string(Image.open(png_path), lang="jpn+eng")
        if not text.strip():
            return fnsku, None

        # 提取描述：去掉 FNSKU 所在行及杂音行
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        desc_lines = []
        for line in lines:
            # 跳过包含文件名FNSKU的行（容忍 FNSKU 中插入空格）
            if fnsku in line or fnsku in re.sub(r"\s+", "", line):
                continue
            # 跳过看起来像FNSKU的行（X00/B0开头），避免OCR误读
            if _is_likely_fnsku(line.split()[0] if line.split() else ""):
                continue
            if re.match(r"^[\*\d\s]+$", line):  # 条码行
                continue
            if "tcpdf" in line.lower():
                continue
            if len(line) < 3:
                continue
            desc_lines.append(line)
        desc = " ".join(desc_lines)
        desc = _clean_desc(desc)
        return fnsku, desc if desc else None
    except Exception:
        return None, None


def process_pdf(pdf_path, png_dir, file_index, dpi):
    """逐页处理单个PDF。

    文本页 → 提取FNSKU + 描述
    图片页 → 转为PNG保存至 png_dir

    返回: (fnsku_results, png_count, error_count)
    """
    fnsku_results = []
    png_count = 0
    error_count = 0

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print("  [ERROR] 无法打开: {}".format(e), file=sys.stderr)
        return [], 0, 1

    basename = os.path.splitext(os.path.basename(pdf_path))[0]

    for page_idx in range(len(doc)):
        page = doc[page_idx]
        text = page.get_text().strip()

        if text:
            # ---- 文本页：提取 FNSKU（按 Y 坐标排序保证视觉顺序）----
            blocks = page.get_text("blocks")
            blocks.sort(key=lambda b: (b[1], b[0]))  # 按 Y, X 排序
            ordered_text = "\n".join(b[4] for b in blocks if b[4].strip())

            matches = _FNSKU_RE.findall(ordered_text)
            loose = False
            if not matches:
                # FNSKU 文本层可能被插入空格（如 "X00 1C40HN3"），宽松匹配重试
                matches = _loose_fnsku_matches(ordered_text)
                loose = bool(matches)
            if not matches:
                continue

            lines = ordered_text.split("\n")
            seen_on_page = set()
            for fnsku in matches:
                if not _is_likely_fnsku(fnsku):
                    continue
                if fnsku in seen_on_page:
                    continue
                seen_on_page.add(fnsku)

                # 定位FNSKU所在行（宽松模式下先对行去空白再比对）
                for i, line in enumerate(lines):
                    if fnsku in (re.sub(r"\s+", "", line) if loose else line):
                        desc = _extract_description(lines, i)
                        if desc:
                            desc = _clean_desc(desc)
                        # blocks 排序后多列排列的 SKU 标签可能描述为空
                        # 回退到原始文本流顺序重新提取
                        if not desc:
                            raw_lines = text.split("\n")
                            for k, raw_line in enumerate(raw_lines):
                                if fnsku in raw_line:
                                    desc = _extract_description(raw_lines, k)
                                    if desc:
                                        desc = _clean_desc(desc)
                                    break
                        # OCR 回退：文本提取仍未拿到描述时，用 OCR 补识别
                        if not desc and _OCR_AVAILABLE:
                            try:
                                pix = page.get_pixmap(matrix=fitz.Matrix(dpi, dpi))
                                tmp_png = "{}_[{}]_{:03d}_ocr.png".format(
                                    basename, file_index, page_idx + 1
                                )
                                tmp_png_path = os.path.join(png_dir, tmp_png)
                                pix.save(tmp_png_path)
                                _fnsku_ocr, desc_ocr = _try_ocr_page(tmp_png_path, pdf_path)
                                if desc_ocr:
                                    desc = desc_ocr
                            except Exception:
                                pass
                        fnsku_results.append((fnsku, desc or ""))
                        break
        else:
            # ---- 图片页：转为 PNG + OCR 识别 ----
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(dpi, dpi))
                png_name = "{}_[{}]_{:03d}.png".format(basename, file_index, page_idx + 1)
                png_path_out = os.path.join(png_dir, png_name)
                pix.save(png_path_out)
                png_count += 1
                # OCR 识别
                fnsku_from_ocr, desc_from_ocr = _try_ocr_page(png_path_out, pdf_path)
                if fnsku_from_ocr:
                    fnsku_results.append((fnsku_from_ocr, desc_from_ocr or ""))
            except Exception as e:
                print("  [ERROR] PNG转换失败 页码{}: {}".format(page_idx + 1, e), file=sys.stderr)
                error_count += 1

    doc.close()
    return fnsku_results, png_count, error_count


def _save_xlsx(output_path, results):
    """将提取结果保存为格式化的 .xlsx 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FNSKU"

    # 样式定义
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    body_font = Font(name="Microsoft YaHei", size=10)
    body_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 表头
    headers = ["#", "FNSKU", "Description"]
    col_widths = [6, 20, 70]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 数据行
    for row_idx, (fnsku, desc) in enumerate(results, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else None

        cell_no = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        cell_no.font = body_font
        cell_no.alignment = Alignment(horizontal="center", vertical="center")
        cell_no.border = thin_border
        if row_fill:
            cell_no.fill = row_fill

        cell_sku = ws.cell(row=row_idx, column=2, value=fnsku)
        cell_sku.font = Font(name="Consolas", size=10)
        cell_sku.alignment = Alignment(vertical="center")
        cell_sku.border = thin_border
        if row_fill:
            cell_sku.fill = row_fill

        cell_desc = ws.cell(row=row_idx, column=3, value=desc)
        cell_desc.font = body_font
        cell_desc.alignment = body_align
        cell_desc.border = thin_border
        if row_fill:
            cell_desc.fill = row_fill

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:C{}".format(len(results) + 1)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="FNSKU和商品描述提取工具 v5.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
 示例:
   python extract_fnsku.py -s "D:\\桌面\\PASU5164708410-629CT 标签"
   python extract_fnsku.py -s ./labels -o results.xlsx -p ocr_pngs -d 3
   python extract_fnsku.py -s ./labels --labels
        """,
    )
    parser.add_argument("--source", "-s", required=True, help="PDF源目录（必填）")
    parser.add_argument(
        "--output",
        "-o",
        default="fnsku_result.xlsx",
        help="输出xlsx路径（默认: fnsku_result.xlsx，相对路径则放在源目录下）",
    )
    parser.add_argument(
        "--png-dir",
        "-p",
        default="ocr_images",
        help="图片页PNG输出目录（默认: ocr_images，相对路径则放在源目录下）",
    )
    parser.add_argument(
        "--dpi", "-d", type=int, default=4, help="PNG分辨率倍数，4≈300DPI（默认: 4）"
    )
    parser.add_argument(
        "--labels",
        "-l",
        action="store_true",
        help="自动生成 50×30mm PDF 条形码标签（需安装 reportlab）",
    )

    args = parser.parse_args()

    source_dir = os.path.abspath(args.source)
    if not os.path.isdir(source_dir):
        print("错误：源目录不存在: {}".format(source_dir), file=sys.stderr)
        sys.exit(1)

    output_xlsx = args.output
    png_dir = args.png_dir

    # 相对路径 → 相对于源目录
    if not os.path.isabs(output_xlsx):
        output_xlsx = os.path.join(source_dir, output_xlsx)
    if not os.path.isabs(png_dir):
        png_dir = os.path.join(source_dir, png_dir)

    os.makedirs(png_dir, exist_ok=True)

    print("=" * 60)
    print("  FNSKU 提取工具 v5.0")
    print("=" * 60)
    print("源目录  : {}".format(source_dir))
    print("输出文件: {}".format(output_xlsx))
    print("PNG目录 : {}".format(png_dir))
    print()

    files = find_sku_files(source_dir)
    print("找到 {} 个符合条件的PDF".format(len(files)))

    if not files:
        print("\n未找到符合条件的PDF文件（文件名需包含 X00 或 B0 且不含 FBA）")
        return

    # ---- Step 1: 逐文件逐页处理 ----
    print("\n[1/3] 逐页处理（文本 → FNSKU / 图片 → PNG）...\n")
    all_results = []
    total_png = 0
    total_errors = 0

    for i, fp in enumerate(files, 1):
        fname = os.path.basename(fp)
        print("[{:03d}/{}] {}".format(i, len(files), fname))

        results, pngs, errors = process_pdf(fp, png_dir, i, args.dpi)
        all_results.extend(results)
        total_png += pngs
        total_errors += errors

        parts = []
        if results:
            parts.append("FNSKU {}条".format(len(results)))
        if pngs:
            parts.append("PNG {}页".format(pngs))
        if errors:
            parts.append("错误 {}处".format(errors))
        if parts:
            print("  -> " + " | ".join(parts))

    # ---- Step 2: 去重 ----
    print("\n[2/3] FNSKU去重（相同FNSKU+描述组合只保留第一条）...")
    seen = set()
    unique_results = []
    for fnsku, desc in all_results:
        key = (fnsku, desc)
        if key not in seen:
            seen.add(key)
            unique_results.append((fnsku, desc))
    print("  原始 {} 条 -> 去重后 {} 条".format(len(all_results), len(unique_results)))

    # ---- Step 3: 保存xlsx ----
    print("\n[3/3] 保存xlsx...")
    xlsx_dir = os.path.dirname(output_xlsx)
    if xlsx_dir:
        os.makedirs(xlsx_dir, exist_ok=True)

    _save_xlsx(output_xlsx, unique_results)

    xlsx_kb = os.path.getsize(output_xlsx) // 1024
    print("  已保存: {} ({} KB)".format(output_xlsx, xlsx_kb))

    # ---- Step 4 (可选): 生成条形码标签 ----
    label_pdf = None
    if args.labels:
        print("\n[4/4] 生成 50×30mm 条形码标签 PDF...")
        try:
            # 确保 scripts/ 目录在 Python 路径中
            _scripts_dir = os.path.dirname(os.path.abspath(__file__))
            if _scripts_dir not in sys.path:
                sys.path.insert(0, _scripts_dir)
            from generate_labels import generate_labels as gen_labels

            result = gen_labels(output_xlsx)
            if result:
                label_pdf, label_count = result
                label_kb = os.path.getsize(label_pdf) // 1024
                print(
                    "  已生成: {} ({} 个标签, {} 页, {} KB)".format(
                        label_pdf, label_count, label_count, label_kb
                    )
                )
        except ImportError:
            print("  [提示] reportlab 未安装，跳过标签生成")
        except Exception as e:
            print("  [错误] 标签生成失败: {}".format(e))

    # ---- 汇总 ----
    print()
    print("=" * 60)
    print("  处理完成")
    print("  PDF文件   : {} 个".format(len(files)))
    print("  FNSKU提取 : {} 条（去重后 {} 条）".format(len(all_results), len(unique_results)))
    print("  图片转PNG : {} 张".format(total_png))
    print("  输出文件  : {}".format(output_xlsx))
    if label_pdf:
        print("  标签PDF   : {}".format(label_pdf))
    if total_errors:
        print("  错误      : {} 处".format(total_errors))
    print("=" * 60)


if __name__ == "__main__":
    main()
