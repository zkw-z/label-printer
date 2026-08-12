# -*- coding: utf-8 -*-
"""
FNSKU 条形码标签生成器 V4（按规格更新）

从 xlsx 文件读取 FNSKU + Description，生成 50×30mm PDF 条形码标签。

标签布局规格：
  标签尺寸  : 50mm × 30mm
  上边距    : 2mm（条码上方留白）
  下边距    : 1mm（描述下方留白）
  左右边距  : 各 2.5mm
  条形码     : Code128，高度 11mm，线宽 0.25mm，宽度自适应 45mm
  FNSKU 文字: Helvetica 9pt，居中，位于条形码下方
  描述文字  : HeiseiMin-W3，8→7→6pt 自适应，左对齐，FNSKU 下方 1.5mm 处开始

⚠️ 此为 V4 稳定版本，非特殊指令不得修改任何逻辑。
"""

import argparse
import os
import sys

from openpyxl import load_workbook
from reportlab.graphics.barcode import code128
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas

# ---- 常量 ----
LABEL_W = 50 * mm
LABEL_H = 30 * mm
MARGIN_TOP = 2 * mm  # 上边距（条码上方留白）
MARGIN_BOTTOM = 1 * mm  # 下边距（描述下方留白）
MARGIN_LR = 2.5 * mm  # 左右边距
BARCODE_H = 11 * mm  # 条形码高度
BARCODE_W_TARGET = 45 * mm  # 条形码目标宽度（标签宽 − 2×边距）
BARCODE_LINE_W = 0.25 * mm  # 条形码线宽
FNSKU_FONT_SIZE = 9  # FNSKU 字号
DESC_FONT_SIZES = [8, 7, 6]  # 描述字号（从大到小尝试）

# CJK 字体（reportlab 内置）
_CJK_FONT = "HeiseiMin-W3"


def _register_cjk_font():
    """注册 CJK 字体，只注册一次"""
    if not hasattr(_register_cjk_font, "_done"):
        pdfmetrics.registerFont(UnicodeCIDFont(_CJK_FONT))
        _register_cjk_font._done = True


def _clean_desc(desc):
    """清理描述中的无意义字符串（如 TCPDF 签名）"""
    if not desc:
        return desc
    # 移除 Powered by TCPDF 相关字符串
    import re

    patterns = [
        r"Powered\s+by\s+TCPDF.*?org[/\w]*",
        r"www\.tcpdf\.org[/\w]*",
        r"tcpdf\.org[/\w]*",
        r"Powered\s+by\s+TCPDF",
    ]
    cleaned = desc
    for p in patterns:
        cleaned = re.sub(p, "", cleaned, flags=re.IGNORECASE)
    # 清理多余空白
    cleaned = " ".join(cleaned.split())
    return cleaned.strip()


def _draw_label(c, fnsku, desc):
    """在当前页绘制一个 50×30mm 标签，从 (0,0) 开始"""
    # 清理描述
    if desc:
        desc = _clean_desc(desc)
    content_w = LABEL_W - 2 * MARGIN_LR  # 45mm
    content_h = LABEL_H - MARGIN_TOP - MARGIN_BOTTOM

    # ---- 1. 条形码 ----
    # 垂直位置：标签高度 − 上边距 − 条形码高度
    barcode_y = LABEL_H - MARGIN_TOP - BARCODE_H

    barcode = code128.Code128(
        fnsku,
        barHeight=BARCODE_H,
        barWidth=BARCODE_LINE_W,
        quiet=0,  # 左右静区设为 0，宽度自适应
    )
    # 水平自适应：缩放画布使条形码宽度 = 目标宽度 45mm
    target_w = BARCODE_W_TARGET  # 45mm
    scale_x = target_w / barcode.width
    cx = MARGIN_LR  # 左对齐起点
    c.saveState()
    c.translate(cx, barcode_y)
    c.scale(scale_x, 1)
    barcode.drawOn(c, 0, 0)
    c.restoreState()

    # ---- 2. FNSKU 文字 ----
    # 垂直位置：条形码底边下方 1mm（视觉间距）
    # 规格公式：条码Y − 0.5mm − 2.5mm
    # 实际实现：barcode_y（条码底边）− 1mm（间距）− 文字视觉高度
    fnsku_visual_h = FNSKU_FONT_SIZE * 1.074  # pt，基线到 bbox 顶边
    fnsku_gap = 1 * mm  # 条形码底边到 FNSKU 文字顶边的间距
    fnsku_y = barcode_y - fnsku_gap - fnsku_visual_h
    c.setFont("Helvetica", FNSKU_FONT_SIZE)
    fnsku_w = pdfmetrics.stringWidth(fnsku, "Helvetica", FNSKU_FONT_SIZE)
    fnsku_x = MARGIN_LR + (content_w - fnsku_w) / 2  # 居中
    c.drawString(fnsku_x, fnsku_y, fnsku)

    # ---- 3. 描述文字（CJK 自适应）----
    # 垂直位置：FNSKU 文字下方 1.5mm 处开始
    desc_area_top = fnsku_y - 1.5 * mm
    desc_area_bottom = MARGIN_BOTTOM
    desc_area_h = desc_area_top - desc_area_bottom

    if not desc:
        return

    def _wrap_lines(text, font, max_w):
        """智能换行：按空格分词（保留标识符编号完整性），
        单词过宽时逐字符拆分。"""
        tokens = text.split(" ")
        lines = []
        current = ""
        for token in tokens:
            if not token:
                continue
            # 先在 token 间加空格（首位不加）
            sep = " " if current else ""
            test = current + sep + token
            w = pdfmetrics.stringWidth(test, font, font_size)
            if w <= max_w:
                current = test
            else:
                # 当前行已有内容 → 保存并另起新行
                if current:
                    lines.append(current)
                    current = token
                else:
                    # token 本身超宽 → 逐字符拆分
                    for ch in token:
                        w2 = pdfmetrics.stringWidth(current + ch, font, font_size)
                        if w2 > max_w and current:
                            lines.append(current)
                            current = ch
                        else:
                            current += ch
        if current:
            lines.append(current)
        return lines

    # 尝试不同字号，找到能放下的
    for font_size in DESC_FONT_SIZES:
        c.setFont(_CJK_FONT, font_size)
        line_spacing = font_size + 1
        max_lines = int(desc_area_h / line_spacing)
        if max_lines < 1:
            continue

        lines = _wrap_lines(desc, _CJK_FONT, content_w)
        if len(lines) <= max_lines:
            for i, line in enumerate(lines[:max_lines]):
                ly = desc_area_top - (i + 1) * line_spacing + font_size * 0.2
                c.drawString(MARGIN_LR, ly, line)
            return

    # 所有字号都放不下：用最小字号，截断
    font_size = DESC_FONT_SIZES[-1]
    c.setFont(_CJK_FONT, font_size)
    line_spacing = font_size + 1
    max_lines = int(desc_area_h / line_spacing)
    lines = _wrap_lines(desc, _CJK_FONT, content_w)
    for i, line in enumerate(lines[:max_lines]):
        ly = desc_area_top - (i + 1) * line_spacing + font_size * 0.2
        c.drawString(MARGIN_LR, ly, line)


def generate_labels(xlsx_path, output_pdf=None):
    """从 xlsx 读取数据，生成 50×30mm PDF 标签（每页一个标签）

    返回: (output_pdf_path, label_count) 或 None
    """
    _register_cjk_font()

    if output_pdf is None:
        base = os.path.splitext(xlsx_path)[0]
        output_pdf = base + "_labels.pdf"

    # 读取 xlsx
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            desc_str = str(row[2]).strip() if row[2] else ""
            items.append((str(row[1]).strip(), desc_str))
    wb.close()

    if not items:
        print("xlsx 中无有效数据", file=sys.stderr)
        return None

    # 页面尺寸 = 标签尺寸 50×30mm
    page_size = (LABEL_W, LABEL_H)

    c = canvas.Canvas(output_pdf, pagesize=page_size)
    c.setTitle("FNSKU Labels")

    for idx, (fnsku, desc) in enumerate(items):
        if idx > 0:
            c.showPage()
        _draw_label(c, fnsku, desc)

    c.save()
    return output_pdf, len(items)


def main():
    parser = argparse.ArgumentParser(
        description="FNSKU 条形码标签生成器 — 从 xlsx 生成 50×30mm PDF 条形码标签（每页一个标签）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
 示例:
   python generate_labels.py -i sku_result.xlsx
   python generate_labels.py -i sku_result.xlsx -o labels.pdf
        """,
    )
    parser.add_argument(
        "--input", "-i", required=True, help="输入 xlsx 路径（由 extract_fnsku.py 生成）"
    )
    parser.add_argument(
        "--output", "-o", default=None, help="输出 PDF 路径（默认: 输入文件同名的 _labels.pdf）"
    )

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print("错误：输入文件不存在: {}".format(args.input), file=sys.stderr)
        sys.exit(1)

    print("=" * 60)
    print("  FNSKU 条形码标签生成器")
    print("=" * 60)
    print("输入文件: {}".format(args.input))
    print("页面尺寸: 50mm × 30mm（每页一个标签）")
    print()

    result = generate_labels(args.input, args.output)

    if result is None:
        print("\n未生成标签（xlsx 中无有效数据）")
        sys.exit(1)

    output_pdf, count = result
    pdf_kb = os.path.getsize(output_pdf) // 1024
    print("  已生成: {}".format(output_pdf))
    print("  标签数: {} 个 ({} 页, {} KB)".format(count, count, pdf_kb))
    print("=" * 60)


if __name__ == "__main__":
    main()
